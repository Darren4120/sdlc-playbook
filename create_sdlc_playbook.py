"""
SDLC Playbook Excel Generator
Creates a professional organization-level SDLC checklist in Excel format
"""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

# Create a new workbook
wb = openpyxl.Workbook()
wb.remove(wb.active)  # Remove default sheet

# Define styles
header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True, size=12)
phase_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
phase_font = Font(color="FFFFFF", bold=True, size=11)
subheader_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
subheader_font = Font(bold=True, size=10)
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# SDLC phases and detailed tasks
sdlc_data = {
    "Planning & Requirements": [
        {
            "category": "Project Initiation",
            "tasks": [
                ("Define project goals and objectives", "Project Manager", "Project Charter", "High"),
                ("Identify stakeholders", "Project Manager", "Stakeholder Register", "High"),
                ("Define project scope", "Business Analyst", "Scope Statement", "High"),
                ("Conduct feasibility study", "Business Analyst", "Feasibility Report", "High"),
                ("Establish project governance", "Project Manager", "Governance Framework", "Medium"),
                ("Create project budget estimate", "Project Manager", "Budget Document", "High"),
            ]
        },
        {
            "category": "Requirements Gathering",
            "tasks": [
                ("Conduct stakeholder interviews", "Business Analyst", "Interview Notes", "High"),
                ("Document business requirements", "Business Analyst", "BRD Document", "High"),
                ("Define functional requirements", "Business Analyst", "FRD Document", "High"),
                ("Define non-functional requirements", "Solution Architect", "NFR Document", "High"),
                ("Create user stories/use cases", "Business Analyst", "User Stories", "High"),
                ("Prioritize requirements (MoSCoW)", "Product Owner", "Prioritized Backlog", "High"),
                ("Requirements validation with stakeholders", "Business Analyst", "Sign-off Document", "High"),
                ("Establish acceptance criteria", "Business Analyst", "Acceptance Criteria", "High"),
            ]
        },
        {
            "category": "Planning & Estimation",
            "tasks": [
                ("Create work breakdown structure (WBS)", "Project Manager", "WBS Document", "High"),
                ("Estimate effort and duration", "Tech Lead", "Estimation Sheet", "High"),
                ("Develop project schedule", "Project Manager", "Project Plan", "High"),
                ("Identify risks and mitigation strategies", "Project Manager", "Risk Register", "High"),
                ("Resource allocation planning", "Project Manager", "Resource Plan", "Medium"),
                ("Define communication plan", "Project Manager", "Communication Plan", "Medium"),
                ("Establish quality metrics", "QA Lead", "Quality Plan", "Medium"),
            ]
        }
    ],
    "Design": [
        {
            "category": "Architecture & Design",
            "tasks": [
                ("Create high-level architecture design", "Solution Architect", "Architecture Document", "High"),
                ("Design system components", "Solution Architect", "Component Diagram", "High"),
                ("Create database schema design", "Database Architect", "ERD/Schema Document", "High"),
                ("Define API specifications", "Solution Architect", "API Documentation", "High"),
                ("Create UI/UX designs and wireframes", "UX Designer", "Design Mockups", "High"),
                ("Design security architecture", "Security Architect", "Security Design Doc", "High"),
                ("Plan integration points", "Solution Architect", "Integration Design", "Medium"),
                ("Create deployment architecture", "DevOps Engineer", "Deployment Diagram", "Medium"),
            ]
        },
        {
            "category": "Design Review",
            "tasks": [
                ("Conduct architecture review", "Architecture Team", "Review Report", "High"),
                ("Perform security design review", "Security Team", "Security Review", "High"),
                ("Review scalability and performance", "Solution Architect", "Performance Review", "Medium"),
                ("Validate design against requirements", "Business Analyst", "Validation Document", "High"),
                ("Obtain design approval", "Stakeholders", "Approval Document", "High"),
            ]
        }
    ],
    "Development": [
        {
            "category": "Environment Setup",
            "tasks": [
                ("Setup development environment", "DevOps Engineer", "Environment Config", "High"),
                ("Configure version control repository", "Tech Lead", "Repository Setup", "High"),
                ("Setup CI/CD pipeline", "DevOps Engineer", "Pipeline Configuration", "High"),
                ("Configure development tools", "Tech Lead", "Tool Configuration", "Medium"),
                ("Setup code quality tools", "Tech Lead", "Quality Tools Config", "Medium"),
            ]
        },
        {
            "category": "Code Development",
            "tasks": [
                ("Implement backend services", "Backend Developer", "Backend Code", "High"),
                ("Implement frontend components", "Frontend Developer", "Frontend Code", "High"),
                ("Implement database scripts", "Database Developer", "DB Scripts", "High"),
                ("Implement API endpoints", "Backend Developer", "API Code", "High"),
                ("Write unit tests", "Developer", "Unit Tests", "High"),
                ("Conduct code reviews", "Tech Lead", "Review Comments", "High"),
                ("Refactor code for quality", "Developer", "Clean Code", "Medium"),
                ("Document code and APIs", "Developer", "Code Documentation", "Medium"),
            ]
        },
        {
            "category": "Code Quality",
            "tasks": [
                ("Run static code analysis", "Developer", "Analysis Report", "High"),
                ("Check code coverage", "Developer", "Coverage Report", "High"),
                ("Address code review findings", "Developer", "Fixed Code", "High"),
                ("Ensure coding standards compliance", "Tech Lead", "Compliance Report", "Medium"),
            ]
        }
    ],
    "Testing": [
        {
            "category": "Test Planning",
            "tasks": [
                ("Create test strategy", "QA Lead", "Test Strategy Doc", "High"),
                ("Develop test plan", "QA Lead", "Test Plan", "High"),
                ("Create test cases", "QA Engineer", "Test Cases", "High"),
                ("Prepare test data", "QA Engineer", "Test Data Sets", "High"),
                ("Setup test environment", "QA Lead", "Test Environment", "High"),
            ]
        },
        {
            "category": "Test Execution",
            "tasks": [
                ("Execute unit testing", "Developer", "Unit Test Results", "High"),
                ("Execute integration testing", "QA Engineer", "Integration Test Results", "High"),
                ("Execute system testing", "QA Engineer", "System Test Results", "High"),
                ("Execute UAT (User Acceptance Testing)", "Business Users", "UAT Results", "High"),
                ("Execute performance testing", "Performance Engineer", "Performance Report", "High"),
                ("Execute security testing", "Security Engineer", "Security Test Report", "High"),
                ("Execute regression testing", "QA Engineer", "Regression Results", "Medium"),
                ("Accessibility testing", "QA Engineer", "Accessibility Report", "Medium"),
            ]
        },
        {
            "category": "Defect Management",
            "tasks": [
                ("Log defects in tracking system", "QA Engineer", "Defect Log", "High"),
                ("Prioritize and assign defects", "QA Lead", "Prioritized Defects", "High"),
                ("Fix critical and high defects", "Developer", "Fixed Code", "High"),
                ("Retest fixed defects", "QA Engineer", "Retest Results", "High"),
                ("Update defect status", "QA Engineer", "Updated Status", "Medium"),
                ("Conduct defect review meetings", "QA Lead", "Meeting Minutes", "Medium"),
            ]
        }
    ],
    "Deployment": [
        {
            "category": "Pre-Deployment",
            "tasks": [
                ("Create deployment plan", "DevOps Engineer", "Deployment Plan", "High"),
                ("Prepare release notes", "Tech Lead", "Release Notes", "High"),
                ("Create rollback plan", "DevOps Engineer", "Rollback Procedures", "High"),
                ("Conduct deployment readiness review", "Release Manager", "Readiness Report", "High"),
                ("Schedule deployment window", "Release Manager", "Deployment Schedule", "High"),
                ("Notify stakeholders", "Project Manager", "Notification Email", "Medium"),
                ("Backup production data", "DBA", "Backup Confirmation", "High"),
            ]
        },
        {
            "category": "Deployment Execution",
            "tasks": [
                ("Deploy to staging environment", "DevOps Engineer", "Staging Deployment", "High"),
                ("Validate staging deployment", "QA Lead", "Validation Results", "High"),
                ("Deploy to production", "DevOps Engineer", "Prod Deployment", "High"),
                ("Execute smoke tests", "QA Engineer", "Smoke Test Results", "High"),
                ("Monitor system health", "DevOps Engineer", "Monitoring Report", "High"),
                ("Verify all services running", "DevOps Engineer", "Service Status", "High"),
            ]
        },
        {
            "category": "Post-Deployment",
            "tasks": [
                ("Conduct post-deployment review", "Release Manager", "Review Report", "Medium"),
                ("Document lessons learned", "Project Manager", "Lessons Learned", "Medium"),
                ("Update deployment documentation", "DevOps Engineer", "Updated Docs", "Medium"),
                ("Confirm with stakeholders", "Project Manager", "Confirmation Email", "High"),
            ]
        }
    ],
    "Maintenance & Support": [
        {
            "category": "Operations Setup",
            "tasks": [
                ("Setup monitoring and alerting", "DevOps Engineer", "Monitoring Config", "High"),
                ("Configure logging", "DevOps Engineer", "Logging Config", "High"),
                ("Create operations runbook", "DevOps Engineer", "Runbook Document", "High"),
                ("Define SLAs and KPIs", "Service Manager", "SLA Document", "High"),
                ("Setup incident management process", "Service Manager", "Incident Process", "High"),
                ("Establish support channels", "Support Manager", "Support Framework", "Medium"),
            ]
        },
        {
            "category": "Ongoing Maintenance",
            "tasks": [
                ("Monitor application performance", "DevOps Engineer", "Performance Metrics", "High"),
                ("Review system logs regularly", "DevOps Engineer", "Log Analysis", "Medium"),
                ("Apply security patches", "DevOps Engineer", "Patch Report", "High"),
                ("Database maintenance", "DBA", "Maintenance Log", "Medium"),
                ("Update system documentation", "Tech Lead", "Updated Docs", "Medium"),
                ("Conduct periodic security audits", "Security Team", "Audit Report", "High"),
            ]
        },
        {
            "category": "Support & Enhancements",
            "tasks": [
                ("Handle user support tickets", "Support Team", "Ticket Resolution", "High"),
                ("Gather user feedback", "Product Owner", "Feedback Log", "Medium"),
                ("Prioritize enhancement requests", "Product Owner", "Prioritized Backlog", "Medium"),
                ("Implement bug fixes", "Developer", "Fixed Code", "High"),
                ("Implement minor enhancements", "Developer", "Enhancement Code", "Medium"),
            ]
        }
    ]
}

# Function to create a detailed phase sheet
def create_phase_sheet(wb, phase_name, categories):
    ws = wb.create_sheet(title=phase_name[:31])  # Excel sheet name limit is 31 chars
    
    # Set column widths
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 30
    
    # Create header row
    headers = ['#', 'Task/Activity', 'Responsible', 'Deliverable', 'Priority', 'Status', 'Target Date', 'Comments']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    row = 2
    task_num = 1
    
    for category_data in categories:
        category = category_data['category']
        
        # Add category row
        cell = ws.cell(row=row, column=1)
        cell.value = ''
        cell.fill = subheader_fill
        cell.border = border
        
        cell = ws.cell(row=row, column=2)
        cell.value = category
        cell.font = subheader_font
        cell.fill = subheader_fill
        cell.border = border
        
        # Merge category cells
        ws.merge_cells(f'B{row}:H{row}')
        for col in range(3, 9):
            ws.cell(row=row, column=col).fill = subheader_fill
            ws.cell(row=row, column=col).border = border
        
        row += 1
        
        # Add tasks
        for task, responsible, deliverable, priority in category_data['tasks']:
            ws.cell(row=row, column=1).value = task_num
            ws.cell(row=row, column=1).alignment = Alignment(horizontal='center')
            ws.cell(row=row, column=2).value = task
            ws.cell(row=row, column=3).value = responsible
            ws.cell(row=row, column=4).value = deliverable
            ws.cell(row=row, column=5).value = priority
            ws.cell(row=row, column=6).value = 'Not Started'
            ws.cell(row=row, column=7).value = ''  # Date field
            ws.cell(row=row, column=8).value = ''  # Comments field
            
            # Apply borders to all cells
            for col in range(1, 9):
                ws.cell(row=row, column=col).border = border
            
            # Set alignment
            ws.cell(row=row, column=2).alignment = Alignment(wrap_text=True, vertical='center')
            ws.cell(row=row, column=4).alignment = Alignment(wrap_text=True, vertical='center')
            
            row += 1
            task_num += 1
    
    # Freeze panes
    ws.freeze_panes = 'A2'
    
    return ws

# Create Overview/Summary sheet
def create_overview_sheet(wb):
    ws = wb.create_sheet(title="Overview", index=0)
    
    # Title
    ws.merge_cells('A1:H1')
    title_cell = ws.cell(row=1, column=1)
    title_cell.value = "SDLC PLAYBOOK - Organization Level Checklist"
    title_cell.font = Font(size=16, bold=True, color="1F4E78")
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 25
    
    # Document info
    ws.merge_cells('A3:B3')
    ws.cell(row=3, column=1).value = "Document Information"
    ws.cell(row=3, column=1).font = Font(bold=True, size=12)
    
    info_data = [
        ('Version:', '1.0'),
        ('Created Date:', datetime.now().strftime('%Y-%m-%d')),
        ('Organization:', '[Your Organization Name]'),
        ('Project Name:', '[Project Name]'),
        ('Project Manager:', '[Name]'),
    ]
    
    row = 4
    for label, value in info_data:
        ws.cell(row=row, column=1).value = label
        ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(row=row, column=2).value = value
        row += 1
    
    # Instructions
    ws.merge_cells(f'A{row+1}:H{row+1}')
    ws.cell(row=row+1, column=1).value = "Instructions for Use"
    ws.cell(row=row+1, column=1).font = Font(bold=True, size=12)
    ws.cell(row=row+1, column=1).fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    
    instructions = [
        "1. This playbook covers all phases of the Software Development Life Cycle (SDLC)",
        "2. Each phase has its own sheet with detailed tasks and activities",
        "3. Update the 'Status' column as tasks are completed (Not Started, In Progress, Completed, Blocked)",
        "4. Assign target dates and responsible parties for each task",
        "5. Use the Comments column to document any issues, dependencies, or notes",
        "6. Priority levels: High (critical), Medium (important), Low (nice-to-have)",
        "7. Regularly review and update the checklist throughout the project lifecycle",
        "8. Customize tasks based on your specific project needs and organizational requirements",
    ]
    
    row += 2
    for instruction in instructions:
        ws.cell(row=row, column=1).value = instruction
        ws.merge_cells(f'A{row}:H{row}')
        row += 1
    
    # Phase summary
    ws.merge_cells(f'A{row+1}:H{row+1}')
    ws.cell(row=row+1, column=1).value = "SDLC Phases Covered"
    ws.cell(row=row+1, column=1).font = Font(bold=True, size=12)
    ws.cell(row=row+1, column=1).fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    
    phase_summary = [
        ("1. Planning & Requirements", "Project initiation, requirements gathering, planning and estimation"),
        ("2. Design", "Architecture design, system design, design review and approval"),
        ("3. Development", "Environment setup, coding, code quality, version control"),
        ("4. Testing", "Test planning, test execution, defect management, UAT"),
        ("5. Deployment", "Pre-deployment prep, deployment execution, post-deployment review"),
        ("6. Maintenance & Support", "Operations setup, ongoing maintenance, support and enhancements"),
    ]
    
    row += 2
    for phase, description in phase_summary:
        ws.cell(row=row, column=1).value = phase
        ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(row=row, column=2).value = description
        ws.merge_cells(f'B{row}:H{row}')
        row += 1
    
    # Set column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 70
    
    return ws

# Create Status Tracking sheet
def create_tracking_sheet(wb):
    ws = wb.create_sheet(title="Status Tracking")
    
    # Title
    ws.merge_cells('A1:F1')
    title_cell = ws.cell(row=1, column=1)
    title_cell.value = "Project Status Summary"
    title_cell.font = Font(size=14, bold=True)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 20
    
    # Headers
    headers = ['Phase', 'Total Tasks', 'Completed', 'In Progress', 'Not Started', '% Complete']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # List phases
    phases = [
        'Planning & Requirements',
        'Design',
        'Development',
        'Testing',
        'Deployment',
        'Maintenance & Support'
    ]
    
    row = 4
    for phase in phases:
        ws.cell(row=row, column=1).value = phase
        ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(row=row, column=1).border = border
        
        for col in range(2, 7):
            ws.cell(row=row, column=col).value = ''
            ws.cell(row=row, column=col).border = border
            ws.cell(row=row, column=col).alignment = Alignment(horizontal='center')
        
        row += 1
    
    # Set column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    
    return ws

# Create all sheets
create_overview_sheet(wb)

for phase_name, categories in sdlc_data.items():
    create_phase_sheet(wb, phase_name, categories)

create_tracking_sheet(wb)

# Save the workbook
output_file = r"c:\VS Code\SDLC PLaybook\SDLC_Playbook_Checklist.xlsx"
wb.save(output_file)

print(f"✓ SDLC Playbook created successfully: {output_file}")
print(f"✓ Total sheets: {len(wb.sheetnames)}")
print(f"✓ Phases covered: {len(sdlc_data)}")
print("\nThe playbook includes:")
print("  - Overview sheet with instructions")
print("  - Detailed checklists for each SDLC phase")
print("  - Status tracking sheet")
print("  - Professional formatting and structure")
